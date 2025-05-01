import glob
import pandas as pd
from openpyxl import load_workbook
from datetime import datetime

current_month = datetime.now().month
prev_month = current_month - 1 if current_month > 1 else 12
two_month_ago = prev_month - 1 if prev_month > 1 else 12


def excel_handler(report_file, sheet_name=None):
    # sheet
    workbook = load_workbook(report_file)
    sheet = workbook[sheet_name] if sheet_name else workbook.active
    old_sheet = workbook[f"{sheet_name}_OLD"] if sheet_name else workbook.active

    # last low
    host_last_row = sum_last_row = sheet.max_row
    while sheet.cell(row=host_last_row, column=2).value is None and host_last_row > 1:
        host_last_row -= 1
    while sheet.cell(row=sum_last_row, column=5).value is None and sum_last_row > 1:
        sum_last_row -= 1

    old_host_last_row = old_sum_last_row = None
    if sheet_name:
        old_host_last_row = old_sum_last_row = old_sheet.max_row
        while workbook[f"{sheet_name}_OLD"].cell(row=old_host_last_row, column=2).value is None and old_host_last_row > 1:
            old_host_last_row -= 1
        while workbook[f"{sheet_name}_OLD"].cell(row=old_sum_last_row, column=5).value is None and old_sum_last_row > 1:
            old_sum_last_row -= 1

    # column
    if report_file != "./statistics_report/教員カラープリンタ印刷統計.xlsx":
        month_col = { 4: "E", 5: "F", 6: "G", 7: "H", 8: "I", 9: "J", 10: "K", 11: "L", 12: "M", 1: "N", 2: "O", 3: "P" }
        prev_month_col = month_col[prev_month]
        two_month_col = month_col[two_month_ago]
        host_col = [ 4, "D" ]
        diff_col = [ 17, "Q" ]
    else: # 教員カラープリンタ印刷統計 の場合
        month_col = { 4: "D", 5: "E", 6: "F", 7: "G", 8: "H", 9: "I", 10: "J", 11: "K", 12: "L", 1: "M", 2: "N", 3: "O" }
        prev_month_col = month_col[prev_month]
        two_month_col = month_col[two_month_ago]
        host_col = [ 3, "C" ]
        diff_col = [ 16, "P" ]

    # header
    excel_head = csv_head = None
    if report_file != "./statistics_report/教員カラープリンタ印刷統計.xlsx":
        if report_file != "./statistics_report/Ricohスキャナ統計.xlsx": # ロビープリンタ印刷統計 教室等モノクロプリンタ印刷統計 の場合
            csv_head = "Printer: B&W"

        else: # Ricohスキャナ統計.xlsx の場合
            if "カラー" in sheet_name:
                excel_head = "AK"
            if "モノクロ" in sheet_name:
                excel_head = "AL"

    else: # 教員カラープリンタ印刷統計 の場合
        if "カラー" in sheet_name:
            csv_head = "Printer: Full Color"
        if "モノクロ" in sheet_name:
            csv_head = "Printer: B&W"

    return report_file, workbook, sheet, old_sheet, host_last_row, sum_last_row, old_host_last_row, old_sum_last_row, prev_month_col, two_month_col, host_col, diff_col, csv_head, excel_head



class commander:
    def __init__(self, report_file, workbook, sheet, old_sheet, host_last_row, sum_last_row, old_host_last_row, old_sum_last_row, prev_month_col, two_month_col, host_col, diff_col, csv_head, excel_head):
        self.report_file = report_file
        self.workbook = workbook
        self.sheet = sheet
        self.old_sheet = old_sheet
        self.host_last_row = host_last_row
        self.sum_last_row = sum_last_row
        self.old_host_last_row = old_host_last_row
        self.old_sum_last_row = old_sum_last_row
        self.prev_month_col = prev_month_col
        self.two_month_col = two_month_col
        self.host_col = host_col
        self.diff_col = diff_col
        self.csv_head = csv_head
        self.excel_head = excel_head


    def fill_in_report(self):
        print(self.report_file)

        if self.report_file != "./statistics_report/Ricohスキャナ統計.xlsx":

            file = glob.glob("./number_report/最新の機器カウンターレポート*")[0]
            reference_file = file.replace(".csv", "_utf8.csv")
            df = pd.read_csv(file, encoding="shift_jis", comment="#", skip_blank_lines=True) # コメント行と空白行を無視
            df.to_csv(reference_file, encoding="utf-8", index=False)
            df_utf8 = pd.read_csv(reference_file, encoding="utf-8")

            for row_idx in range(3, self.host_last_row + 1):
                host_name = self.sheet[f"{self.host_col[1]}{row_idx}"].value

                if host_name:
                    if host_name in df_utf8["Host Name"].values:

                        value = df_utf8[df_utf8["Host Name"] == host_name][self.csv_head].values[0]
                    else:
                        print(f"{host_name} が見つかりませんでした")

                    self.sheet[f"{self.prev_month_col}{row_idx}"].value = value

        else: # Ricohスキャナ統計.xlsxの場合
            reference_file = glob.glob("./number_report/機能×カラー別集計レポート*.xlsx")[0]
            reference_sheet = excel_handler(reference_file)[3]

            for row_idx in range(3, self.host_last_row + 1):
                host_name = self.sheet[f"{self.host_col[1]}{row_idx}"].value

                if host_name:
                    for row in range(7, 84):
                        find = 0
                        reference_host_name = reference_sheet[f"B{row}"].value
                        
                        if host_name == reference_host_name:
                            value = reference_sheet[f"{self.excel_head}{row}"].value
                            find = 1
                            break
                            
                    if find == 0:
                        print(f"{host_name} が見つかりませんでした")

                    self.sheet[f"{self.prev_month_col}{row_idx}"].value = value

        if prev_month != 4:
            for row_idx in range(3, self.host_last_row + 1):

                prev_value = self.sheet[f"{self.prev_month_col}{row_idx}"].value

                if self.report_file != "./statistics_report/Ricohスキャナ統計.xlsx":

                    two_value = self.sheet[f"{self.two_month_col}{row_idx}"].value
                    diff_value = (prev_value if prev_value else 0) - (two_value if two_value else 0)
                    self.sheet[f"{self.diff_col[1]}{row_idx}"].value = diff_value

                else:
                    self.sheet[f"Q{row_idx}"].value = prev_value

        else:
            host_row_map = {self.old_sheet[f"{self.host_col[1]}{row}"].value: row for row in range(3, self.old_sum_last_row + 1)}

            for row_idx in range(3, self.host_last_row + 1):
                
                prev_value = self.sheet[f"{self.prev_month_col}{row_idx}"].value          

                if self.report_file != "./statistics_report/Ricohスキャナ統計.xlsx":
                    host_name = self.sheet[f"{self.host_col[1]}{row_idx}"].value

                    if host_name in host_row_map:

                        old_row_idx = host_row_map[host_name]
                        two_value = self.old_sheet[f"{self.two_month_col}{old_row_idx}"].value
                        diff_value = (prev_value if prev_value else 0) - (two_value if two_value else 0)
                        self.sheet[f"{self.diff_col[1]}{row_idx}"].value = diff_value

                    else:
                        self.sheet[f"{self.diff_col[1]}{row_idx}"].value = prev_value

                else:
                    self.sheet[f"Q{row_idx}"].value = prev_value

        self.workbook.save(self.report_file)
        print("----------")



    def gen_text(self):

        prev_sum_value = sum(self.sheet[cell].value for cell in [f"{self.prev_month_col}{i}" for i in range(3, self.host_last_row + 1)])
        self.sheet[f"{self.prev_month_col}{self.sum_last_row -1}"].value = prev_sum_value

        if prev_month != 4:
            two_sum_value = sum(self.sheet[cell].value for cell in [f"{self.two_month_col}{i}" for i in range(3, self.host_last_row + 1)])
        else:
            two_sum_value = sum(self.old_sheet[cell].value for cell in [f"{self.two_month_col}{i}" for i in range(3, self.old_host_last_row + 1)])

        if self.report_file != "./statistics_report/Ricohスキャナ統計.xlsx":
            diff_value = (prev_sum_value if prev_sum_value else 0) - (two_sum_value if two_sum_value else 0)
            self.sheet[f"{self.prev_month_col}{self.sum_last_row}"].value = diff_value

        else:
            self.sheet[f"{self.prev_month_col}{self.sum_last_row}"].value = prev_sum_value

        self.workbook.save(self.report_file)

        if prev_sum_value > two_sum_value:
            word = "増加"
            if prev_sum_value - two_sum_value >= 20000:
                word = "大幅に増加"
        elif prev_sum_value < two_sum_value:
            word = "減少"
            if two_sum_value - prev_sum_value >= 20000:
                word = "大幅に減少"
            elif prev_sum_value == 0:
                word = "減少（利用なし）"
        else:
            word = "変化なし"
            if prev_sum_value == 0:
                word = "変化なし（利用なし）"

        max_value = 0
        max_row_num = None

        for row in self.sheet.iter_rows(min_row=2, max_row=self.sum_last_row, min_col=self.diff_col[0], max_col=self.diff_col[0]):
            value = row[0].value

            if value and isinstance(value, (int)):
                if value > max_value:
                    max_value = value
                    max_row_num = row[0].row

        if max_row_num:
            if self.report_file != "./statistics_report/ロビープリンタ印刷統計.xlsx":
                place = f"{self.sheet[f'A{max_row_num}'].value} {self.sheet[f'B{max_row_num}'].value}"
            
            else: # ロビープリンタ印刷統計.xlsxの場合はA,B,C列を取得
                place = f"{self.sheet[f'A{max_row_num}'].value} {self.sheet[f'B{max_row_num}'].value} {self.sheet[f'C{max_row_num}'].value}"
        else:
            place = "該当なし"

        return word, place, max_value



# create instance
commander(*excel_handler("./statistics_report/教室等モノクロプリンタ印刷統計.xlsx", "ALL")).fill_in_report()
commander(*excel_handler("./statistics_report/教員カラープリンタ印刷統計.xlsx", "カラー")).fill_in_report()
commander(*excel_handler("./statistics_report/教員カラープリンタ印刷統計.xlsx", "モノクロ")).fill_in_report()
commander(*excel_handler("./statistics_report/ロビープリンタ印刷統計.xlsx", "ALL")).fill_in_report()
commander(*excel_handler("./statistics_report/Ricohスキャナ統計.xlsx", "教室カラー")).fill_in_report()
commander(*excel_handler("./statistics_report/Ricohスキャナ統計.xlsx", "教室モノクロ")).fill_in_report()
commander(*excel_handler("./statistics_report/Ricohスキャナ統計.xlsx", "ロビーカラー")).fill_in_report()
commander(*excel_handler("./statistics_report/Ricohスキャナ統計.xlsx", "ロビーモノクロ")).fill_in_report()

# create mail report
class_print = commander(*excel_handler("./statistics_report/教室等モノクロプリンタ印刷統計.xlsx", "ALL")).gen_text()
teacher_print_color = commander(*excel_handler("./statistics_report/教員カラープリンタ印刷統計.xlsx", "カラー")).gen_text()
teacher_print_mono = commander(*excel_handler("./statistics_report/教員カラープリンタ印刷統計.xlsx", "モノクロ")).gen_text()
lobby_print = commander(*excel_handler("./statistics_report/ロビープリンタ印刷統計.xlsx", "ALL")).gen_text()
class_scan_color = commander(*excel_handler("./statistics_report/Ricohスキャナ統計.xlsx", "教室カラー")).gen_text()
class_scan_mono = commander(*excel_handler("./statistics_report/Ricohスキャナ統計.xlsx", "教室モノクロ")).gen_text()
lobby_scan_color = commander(*excel_handler("./statistics_report/Ricohスキャナ統計.xlsx", "ロビーカラー")).gen_text()
lobby_scan_mono = commander(*excel_handler("./statistics_report/Ricohスキャナ統計.xlsx", "ロビーモノクロ")).gen_text()

report_text = f"2025年{prev_month}月 利用統計レポート\n\
{prev_month}月分の利用統計レポートを作成しましたのでご確認ください。\n\
----------\n\
・プリンタ，スキャナ\n\
当月は, 教室プリンタは{class_print[0]}, \
教員室プリンタは, カラーが{teacher_print_color[0]}, 白黒が{teacher_print_mono[0]}。\
ロビープリンタは{lobby_print[0]}, \
教室プリンタのスキャナは, カラーが{class_scan_color[0]}, 白黒が{class_scan_mono[0]}。\
ロビープリンタのスキャナは, カラーが{lobby_scan_color[0]}, 白黒が{lobby_scan_mono[0]}。\n\
\
特に目立った利用は，\
教室プリンタ ( {class_print[1]} → {class_print[2]} 枚 ), \
教員室プリンタ ( カラー: {teacher_print_color[1]} → {teacher_print_color[2]} 枚，白黒: {teacher_print_mono[1]} → {teacher_print_mono[2]} 枚 ), \
ロビープリンタ ( {lobby_print[1]} → {lobby_print[2]} 枚 ), \
教室プリンタのスキャナ ( カラー: {class_scan_color[1]} → {class_scan_color[2]} 枚，白黒: {class_scan_mono[1]} → {class_scan_mono[2]} 枚 ), \
ロビープリンタのスキャナ ( カラー: {lobby_scan_color[1]} → {lobby_scan_color[2]} 枚，白黒: {lobby_scan_mono[1]} → {lobby_scan_mono[2]} 枚 )だった。\n\
----------"

print(report_text)
